from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import pandas as pd
import os
from datetime import datetime
import matplotlib.pyplot as plt
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

# Data storage (in a real app, you would use a database)
pidum_data = []
pidsus_data = []

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/input_pidum', methods=['GET', 'POST'])
def input_pidum():
    if request.method == 'POST':
        try:
            # Get form data
            no = request.form['no']
            periode = request.form['periode']
            tanggal = request.form['tanggal']
            jenis_perkara = request.form['jenis_perkara']
            pra_penututan = request.form['pra_penututan']
            penuntutan = request.form['penuntutan']
            upaya_hukum = request.form['upaya_hukum']
            
            # Add to data list
            pidum_data.append({
                'NO': no,
                'PERIODE': periode,
                'TANGGAL': tanggal,
                'JENIS PERKARA': jenis_perkara,
                'PRA PENUTUTAN': pra_penututan,
                'PENUNTUTAN': penuntutan,
                'UPAYA HUKUM': upaya_hukum
            })
            
            flash('Data PIDUM berhasil ditambahkan!', 'success')
            return redirect(url_for('input_pidum'))
        except Exception as e:
            flash(f'Terjadi kesalahan: {str(e)}', 'error')
    
    return render_template('input_pidum.html', data=pidum_data)

@app.route('/input_pidsus', methods=['GET', 'POST'])
def input_pidsus():
    if request.method == 'POST':
        try:
            # Get form data
            no = request.form['no']
            periode = request.form['periode']
            tanggal = request.form['tanggal']
            jenis_perkara = request.form['jenis_perkara']
            penyidikan = request.form['penyidikan']
            penuntutan = request.form['penuntutan']
            keterangan = request.form['keterangan']
            
            # Add to data list
            pidsus_data.append({
                'NO': no,
                'PERIODE': periode,
                'TANGGAL': tanggal,
                'JENIS PERKARA': jenis_perkara,
                'PENYIDIKAN': penyidikan,
                'PENUNTUTAN': penuntutan,
                'KETERANGAN': keterangan
            })
            
            flash('Data PIDSUS berhasil ditambahkan!', 'success')
            return redirect(url_for('input_pidsus'))
        except Exception as e:
            flash(f'Terjadi kesalahan: {str(e)}', 'error')
    
    return render_template('input_pidsus.html', data=pidsus_data)

@app.route('/view_pidum')
def view_pidum():
    return render_template('view_pidum.html', data=pidum_data)

@app.route('/view_pidsus')
def view_pidsus():
    return render_template('view_pidsus.html', data=pidsus_data)

@app.route('/export_pidum_excel')
def export_pidum_excel():
    if not pidum_data:
        flash('Tidak ada data PIDUM untuk diekspor', 'warning')
        return redirect(url_for('view_pidum'))
    
    # Create DataFrame
    df = pd.DataFrame(pidum_data)
    
    # Create Excel file
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Data PIDUM', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Data PIDUM']
        
        # Style the header
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:  # First row
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column width
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"data_pidum_{timestamp}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/export_pidsus_excel')
def export_pidsus_excel():
    if not pidsus_data:
        flash('Tidak ada data PIDSUS untuk diekspor', 'warning')
        return redirect(url_for('view_pidsus'))
    
    # Create DataFrame
    df = pd.DataFrame(pidsus_data)
    
    # Create Excel file
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Data PIDSUS', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Data PIDSUS']
        
        # Style the header
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:  # First row
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column width
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"data_pidsus_{timestamp}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/pidum_charts')
def pidum_charts():
    if not pidum_data:
        flash('Tidak ada data PIDUM untuk ditampilkan dalam grafik', 'warning')
        return redirect(url_for('view_pidum'))
    
    # Create charts
    charts = {}
    
    # Chart by Jenis Perkara
    df = pd.DataFrame(pidum_data)
    jenis_perkara_counts = df['JENIS PERKARA'].value_counts()
    
    plt.figure(figsize=(10, 6))
    jenis_perkara_counts.plot(kind='bar', color='skyblue')
    plt.title('Jumlah Kasus Berdasarkan Jenis Perkara (PIDUM)')
    plt.xlabel('Jenis Perkara')
    plt.ylabel('Jumlah Kasus')
    plt.xticks(rotation=45)
    plt.tight_layout()
    
    img = io.BytesIO()
    plt.savefig(img, format='png')
    img.seek(0)
    charts['jenis_perkara'] = base64.b64encode(img.getvalue()).decode('utf-8')
    plt.close()
    
    # Chart by Upaya Hukum
    upaya_hukum_counts = df['UPAYA HUKUM'].value_counts()
    
    plt.figure(figsize=(8, 6))
    upaya_hukum_counts.plot(kind='pie', autopct='%1.1f%%', colors=['lightcoral', 'lightgreen'])
    plt.title('Distribusi Upaya Hukum (PIDUM)')
    plt.tight_layout()
    
    img = io.BytesIO()
    plt.savefig(img, format='png')
    img.seek(0)
    charts['upaya_hukum'] = base64.b64encode(img.getvalue()).decode('utf-8')
    plt.close()
    
    return render_template('pidum_charts.html', charts=charts)

@app.route('/pidsus_charts')
def pidsus_charts():
    if not pidsus_data:
        flash('Tidak ada data PIDSUS untuk ditampilkan dalam grafik', 'warning')
        return redirect(url_for('view_pidsus'))
    
    # Create charts
    charts = {}
    
    # Chart by Jenis Perkara
    df = pd.DataFrame(pidsus_data)
    jenis_perkara_counts = df['JENIS PERKARA'].value_counts()
    
    plt.figure(figsize=(10, 6))
    jenis_perkara_counts.plot(kind='bar', color='skyblue')
    plt.title('Jumlah Kasus Berdasarkan Jenis Perkara (PIDSUS)')
    plt.xlabel('Jenis Perkara')
    plt.ylabel('Jumlah Kasus')
    plt.xticks(rotation=45)
    plt.tight_layout()
    
    img = io.BytesIO()
    plt.savefig(img, format='png')
    img.seek(0)
    charts['jenis_perkara'] = base64.b64encode(img.getvalue()).decode('utf-8')
    plt.close()
    
    # Chart by Penyidikan
    penyidikan_counts = df['PENYIDIKAN'].value_counts()
    
    plt.figure(figsize=(8, 6))
    penyidikan_counts.plot(kind='pie', autopct='%1.1f%%', colors=['lightcoral', 'lightgreen'])
    plt.title('Distribusi Penyidikan (PIDSUS)')
    plt.tight_layout()
    
    img = io.BytesIO()
    plt.savefig(img, format='png')
    img.seek(0)
    charts['penyidikan'] = base64.b64encode(img.getvalue()).decode('utf-8')
    plt.close()
    
    return render_template('pidsus_charts.html', charts=charts)

if __name__ == '__main__':
    app.run(debug=True, port=5001)